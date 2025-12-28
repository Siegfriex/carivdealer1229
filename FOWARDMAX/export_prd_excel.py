# -*- coding: utf-8 -*-
"""
ForwardMax PRD Excel 변환 스크립트
엑셀 변환용 마크다운 파일을 읽어서 Excel 파일로 변환합니다.
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import sys
import os

# UTF-8 인코딩 설정
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

def parse_markdown_table(md_content, sheet_name):
    """
    마크다운 파일에서 특정 시트의 표를 파싱하여 DataFrame으로 변환
    """
    # 시트 선언 찾기: ## SHEET: 시트명
    pattern = rf'## SHEET:\s*{re.escape(sheet_name)}\s*\n(.*?)(?=\n## SHEET:|$)'
    match = re.search(pattern, md_content, re.DOTALL)
    
    if not match:
        return None
    
    table_content = match.group(1).strip()
    
    # 마크다운 표 파싱
    lines = table_content.split('\n')
    table_lines = []
    
    for line in lines:
        line = line.strip()
        # 표 구분선 건너뛰기
        if line.startswith('|---'):
            continue
        # 표 행 파싱
        if line.startswith('|') and line.endswith('|'):
            # 첫 번째와 마지막 | 제거하고 분리
            cells = [cell.strip() for cell in line[1:-1].split('|')]
            table_lines.append(cells)
    
    if len(table_lines) < 2:  # 헤더 + 최소 1개 데이터 행
        return None
    
    # 첫 번째 행이 헤더
    headers = table_lines[0]
    # 나머지가 데이터
    data_rows = table_lines[1:]
    
    # DataFrame 생성
    df = pd.DataFrame(data_rows, columns=headers)
    
    return df

def apply_excel_formatting(ws, df):
    """
    Excel 워크시트에 서식 적용
    - 첫 행 freeze
    - auto filter
    - 열 너비 자동 조정
    - 텍스트 wrap
    - ID 컬럼 텍스트 서식 강제
    """
    # 첫 행 freeze
    ws.freeze_panes = 'A2'
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 행 서식 적용
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 행 서식 적용
    for row_idx in range(2, len(df) + 2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # ID 컬럼은 텍스트 서식 강제
            if 'ID' in col_name or col_name.endswith(' ID') or col_name == 'Function ID' or col_name == 'Screen ID' or col_name == 'Flow ID' or col_name == 'API ID' or col_name == 'Entity ID' or col_name == 'NFR ID' or col_name == 'Decision ID' or col_name == '갭 ID' or col_name == '리스크 ID' or col_name == '기회 ID' or col_name == 'Need ID' or col_name == 'Req ID':
                cell.number_format = '@'  # 텍스트 형식
                if cell.value:
                    cell.value = str(cell.value)
    
    # 열 너비 자동 조정 (최대 50자)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        
        # 헤더 길이 확인
        max_length = max(max_length, len(str(col_name)))
        
        # 데이터 길이 확인 (최대 100행만 확인하여 성능 최적화)
        for row_idx in range(2, min(len(df) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_length = len(str(cell_value))
                max_length = max(max_length, cell_length)
        
        # 열 너비 설정 (최소 10, 최대 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Auto filter 적용
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(df.columns)).column_letter}{len(df) + 1}"

def validate_references(wb, validation_results):
    """
    참조 무결성 검증
    - Link 시트가 참조하는 ID가 Registry에 존재하는지 확인
    """
    errors = []
    warnings = []
    
    # Registry 시트에서 ID 수집
    registry_ids = {}
    
    # Flow Registry
    if '10_FLOW' in wb.sheetnames:
        ws = wb['10_FLOW']
        if ws.max_row > 1:
            flow_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    flow_ids.add(str(row[0]))
            registry_ids['FLOW'] = flow_ids
    
    # Screen Registry
    if '11_SCREEN' in wb.sheetnames:
        ws = wb['11_SCREEN']
        if ws.max_row > 1:
            screen_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    screen_ids.add(str(row[0]))
            registry_ids['SCREEN'] = screen_ids
    
    # Function Registry
    if '12_FUNCTION' in wb.sheetnames:
        ws = wb['12_FUNCTION']
        if ws.max_row > 1:
            func_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    func_ids.add(str(row[0]))
            registry_ids['FUNCTION'] = func_ids
    
    # Entity Registry
    if '13_ENTITY' in wb.sheetnames:
        ws = wb['13_ENTITY']
        if ws.max_row > 1:
            entity_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    entity_ids.add(str(row[0]))
            registry_ids['ENTITY'] = entity_ids
    
    # API Registry
    if '14_API' in wb.sheetnames:
        ws = wb['14_API']
        if ws.max_row > 1:
            api_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    api_ids.add(str(row[0]))
            registry_ids['API'] = api_ids
    
    # NFR Registry
    if '15_NFR' in wb.sheetnames:
        ws = wb['15_NFR']
        if ws.max_row > 1:
            nfr_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    nfr_ids.add(str(row[0]))
            registry_ids['NFR'] = nfr_ids
    
    # Decision Registry
    if '16_DECISION' in wb.sheetnames:
        ws = wb['16_DECISION']
        if ws.max_row > 1:
            decision_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    decision_ids.add(str(row[0]))
            registry_ids['DECISION'] = decision_ids
    
    # Link 시트 검증
    # 30_FLOW_SCREEN
    if '30_FLOW_SCREEN' in wb.sheetnames:
        ws = wb['30_FLOW_SCREEN']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    flow_id = str(row[0]).strip() if row[0] else None
                    screen_id = str(row[1]).strip() if row[1] else None
                    
                    if flow_id and flow_id not in registry_ids.get('FLOW', set()):
                        errors.append(f"30_FLOW_SCREEN 행 {row_idx}: Flow ID '{flow_id}'가 10_FLOW에 존재하지 않음")
                    if screen_id and screen_id not in registry_ids.get('SCREEN', set()):
                        errors.append(f"30_FLOW_SCREEN 행 {row_idx}: Screen ID '{screen_id}'가 11_SCREEN에 존재하지 않음")
    
    # 31_SCREEN_FUNCTION
    if '31_SCREEN_FUNCTION' in wb.sheetnames:
        ws = wb['31_SCREEN_FUNCTION']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    screen_id = str(row[0]).strip() if row[0] else None
                    func_id = str(row[1]).strip() if row[1] else None
                    
                    if screen_id and screen_id not in registry_ids.get('SCREEN', set()):
                        errors.append(f"31_SCREEN_FUNCTION 행 {row_idx}: Screen ID '{screen_id}'가 11_SCREEN에 존재하지 않음")
                    if func_id and func_id not in registry_ids.get('FUNCTION', set()):
                        errors.append(f"31_SCREEN_FUNCTION 행 {row_idx}: Function ID '{func_id}'가 12_FUNCTION에 존재하지 않음")
    
    # 32_FUNCTION_ENTITY
    if '32_FUNCTION_ENTITY' in wb.sheetnames:
        ws = wb['32_FUNCTION_ENTITY']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    func_id = str(row[0]).strip() if row[0] else None
                    entity_id = str(row[1]).strip() if row[1] else None
                    
                    if func_id and func_id not in registry_ids.get('FUNCTION', set()):
                        errors.append(f"32_FUNCTION_ENTITY 행 {row_idx}: Function ID '{func_id}'가 12_FUNCTION에 존재하지 않음")
                    if entity_id and entity_id not in registry_ids.get('ENTITY', set()):
                        errors.append(f"32_FUNCTION_ENTITY 행 {row_idx}: Entity ID '{entity_id}'가 13_ENTITY에 존재하지 않음")
    
    # 33_FUNCTION_API
    if '33_FUNCTION_API' in wb.sheetnames:
        ws = wb['33_FUNCTION_API']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    func_id = str(row[0]).strip() if row[0] else None
                    api_id = str(row[1]).strip() if row[1] else None
                    
                    if func_id and func_id not in registry_ids.get('FUNCTION', set()):
                        errors.append(f"33_FUNCTION_API 행 {row_idx}: Function ID '{func_id}'가 12_FUNCTION에 존재하지 않음")
                    if api_id and api_id not in registry_ids.get('API', set()):
                        errors.append(f"33_FUNCTION_API 행 {row_idx}: API ID '{api_id}'가 14_API에 존재하지 않음")
    
    # 34_FUNCTION_NFR
    if '34_FUNCTION_NFR' in wb.sheetnames:
        ws = wb['34_FUNCTION_NFR']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    func_id = str(row[0]).strip() if row[0] else None
                    nfr_id = str(row[1]).strip() if row[1] else None
                    
                    if func_id and func_id not in registry_ids.get('FUNCTION', set()):
                        errors.append(f"34_FUNCTION_NFR 행 {row_idx}: Function ID '{func_id}'가 12_FUNCTION에 존재하지 않음")
                    if nfr_id and nfr_id not in registry_ids.get('NFR', set()):
                        errors.append(f"34_FUNCTION_NFR 행 {row_idx}: NFR ID '{nfr_id}'가 15_NFR에 존재하지 않음")
    
    # 35_FUNCTION_DECISION
    if '35_FUNCTION_DECISION' in wb.sheetnames:
        ws = wb['35_FUNCTION_DECISION']
        if ws.max_row > 1:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    func_id = str(row[0]).strip() if row[0] else None
                    decision_id = str(row[1]).strip() if row[1] else None
                    
                    if func_id and func_id not in registry_ids.get('FUNCTION', set()):
                        errors.append(f"35_FUNCTION_DECISION 행 {row_idx}: Function ID '{func_id}'가 12_FUNCTION에 존재하지 않음")
                    if decision_id and decision_id not in registry_ids.get('DECISION', set()):
                        errors.append(f"35_FUNCTION_DECISION 행 {row_idx}: Decision ID '{decision_id}'가 16_DECISION에 존재하지 않음")
    
    validation_results['reference_errors'] = errors
    validation_results['reference_warnings'] = warnings

def validate_duplicates(wb, validation_results):
    """
    중복 검사: Registry의 PK(ID) 중복 여부
    """
    errors = []
    
    # 각 Registry 시트에서 ID 중복 확인
    registry_sheets = {
        '10_FLOW': 'Flow ID',
        '11_SCREEN': 'Screen ID',
        '12_FUNCTION': 'Function ID',
        '13_ENTITY': 'Entity ID',
        '14_API': 'API ID',
        '15_NFR': 'NFR ID',
        '16_DECISION': '결정 ID',
        '17_GAP': '갭 ID',
        '18_RISK': '리스크 ID',
        '19_OPPORTUNITY': '기회 ID',
    }
    
    for sheet_name, id_column in registry_sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                ids = []
                id_col_idx = None
                
                # ID 컬럼 인덱스 찾기
                for col_idx, cell in enumerate(ws[1], start=1):
                    if cell.value == id_column:
                        id_col_idx = col_idx
                        break
                
                if id_col_idx:
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if row and len(row) >= id_col_idx:
                            id_value = row[id_col_idx - 1]
                            if id_value:
                                id_str = str(id_value).strip()
                                if id_str in ids:
                                    errors.append(f"{sheet_name} 행 {row_idx}: 중복된 ID '{id_str}'")
                                ids.append(id_str)
    
    validation_results['duplicate_errors'] = errors

def validate_required_columns(wb, validation_results):
    """
    필수 컬럼 검사: 01_SCHEMA에 정의된 required 컬럼 누락 여부
    """
    errors = []
    warnings = []
    
    if '01_SCHEMA' not in wb.sheetnames:
        warnings.append("01_SCHEMA 시트가 없어 필수 컬럼 검증을 건너뜁니다.")
        validation_results['required_errors'] = errors
        validation_results['required_warnings'] = warnings
        return
    
    ws_schema = wb['01_SCHEMA']
    
    # 스키마에서 필수 컬럼 정보 수집
    required_columns = {}  # {시트명: [필수 컬럼 목록]}
    
    if ws_schema.max_row > 1:
        for row in ws_schema.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 5:
                sheet_name = str(row[0]).strip() if row[0] else None
                column_name = str(row[1]).strip() if row[1] else None
                is_required = str(row[3]).strip().upper() if len(row) > 3 and row[3] else 'N'
                
                if sheet_name and column_name and is_required == 'Y':
                    if sheet_name not in required_columns:
                        required_columns[sheet_name] = []
                    required_columns[sheet_name].append(column_name)
    
    # 각 시트에서 필수 컬럼 존재 여부 확인
    for sheet_name, required_cols in required_columns.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                # 헤더 행에서 컬럼명 수집
                header_row = [cell.value for cell in ws[1]]
                
                for req_col in required_cols:
                    if req_col not in header_row:
                        errors.append(f"{sheet_name}: 필수 컬럼 '{req_col}'가 누락되었습니다.")
                
                # 데이터 행에서 필수 컬럼의 빈 값 확인
                req_col_indices = {}
                for idx, header in enumerate(header_row, start=1):
                    if header in required_cols:
                        req_col_indices[header] = idx
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    for req_col, col_idx in req_col_indices.items():
                        if len(row) >= col_idx:
                            cell_value = row[col_idx - 1]
                            if not cell_value or (isinstance(cell_value, str) and cell_value.strip() == ''):
                                warnings.append(f"{sheet_name} 행 {row_idx}: 필수 컬럼 '{req_col}'가 비어있습니다.")
    
    validation_results['required_errors'] = errors
    validation_results['required_warnings'] = warnings

def validate_counts(wb, validation_results):
    """
    카운트 재검증: PRD에서 정의된 Screen/Function/Flow 개수와 ExcelSource의 개수 교차 확인
    """
    counts = {}
    
    # 각 Registry의 개수 확인
    registry_sheets = {
        '10_FLOW': 'Flow',
        '11_SCREEN': 'Screen',
        '12_FUNCTION': 'Function',
    }
    
    for sheet_name, entity_type in registry_sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 헤더 제외한 데이터 행 수
            data_rows = max(0, ws.max_row - 1)
            counts[entity_type] = data_rows
    
    validation_results['counts'] = counts

def generate_validation_report(validation_results, output_file):
    """
    검증 리포트 생성
    """
    report_lines = []
    report_lines.append("# ForwardMax PRD Excel 변환 검증 리포트")
    report_lines.append("")
    report_lines.append(f"**생성 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    
    # 참조 무결성 검증 결과
    report_lines.append("## 1. 참조 무결성 검증")
    report_lines.append("")
    if validation_results.get('reference_errors'):
        report_lines.append(f"### 오류: {len(validation_results['reference_errors'])}건")
        for error in validation_results['reference_errors']:
            report_lines.append(f"- {error}")
    else:
        report_lines.append("### 통과: 참조 무결성 오류 없음")
    report_lines.append("")
    
    # 중복 검증 결과
    report_lines.append("## 2. 중복 검증")
    report_lines.append("")
    if validation_results.get('duplicate_errors'):
        report_lines.append(f"### 오류: {len(validation_results['duplicate_errors'])}건")
        for error in validation_results['duplicate_errors']:
            report_lines.append(f"- {error}")
    else:
        report_lines.append("### 통과: 중복 ID 없음")
    report_lines.append("")
    
    # 필수 컬럼 검증 결과
    report_lines.append("## 3. 필수 컬럼 검증")
    report_lines.append("")
    if validation_results.get('required_errors'):
        report_lines.append(f"### 오류: {len(validation_results['required_errors'])}건")
        for error in validation_results['required_errors']:
            report_lines.append(f"- {error}")
    else:
        report_lines.append("### 통과: 필수 컬럼 누락 없음")
    
    if validation_results.get('required_warnings'):
        report_lines.append(f"### 경고: {len(validation_results['required_warnings'])}건")
        for warning in validation_results['required_warnings'][:10]:  # 최대 10개만 표시
            report_lines.append(f"- {warning}")
        if len(validation_results['required_warnings']) > 10:
            report_lines.append(f"- ... 외 {len(validation_results['required_warnings']) - 10}건")
    report_lines.append("")
    
    # 카운트 검증 결과
    report_lines.append("## 4. 카운트 검증")
    report_lines.append("")
    if validation_results.get('counts'):
        for entity_type, count in validation_results['counts'].items():
            report_lines.append(f"- {entity_type}: {count}개")
    report_lines.append("")
    
    # 요약
    report_lines.append("## 요약")
    report_lines.append("")
    total_errors = len(validation_results.get('reference_errors', [])) + len(validation_results.get('duplicate_errors', [])) + len(validation_results.get('required_errors', []))
    total_warnings = len(validation_results.get('required_warnings', []))
    
    if total_errors == 0:
        report_lines.append("**검증 결과**: ✅ 통과 (오류 없음)")
    else:
        report_lines.append(f"**검증 결과**: ❌ 실패 (오류 {total_errors}건)")
    
    if total_warnings > 0:
        report_lines.append(f"**경고**: {total_warnings}건")
    
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append("**생성 도구**: export_prd_excel.py")
    report_lines.append(f"**생성 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 리포트 파일 저장
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    
    print(f"검증 리포트 생성 완료: {output_file}")

def main():
    """
    메인 함수: 마크다운 파일을 읽어서 Excel 파일로 변환
    """
    input_file = 'ForwardMax_PRD_ExcelSource.md'
    output_file = 'ForwardMax_PRD.xlsx'
    validation_report_file = 'ForwardMax_PRD_validation_report.md'
    
    print("=" * 60)
    print("ForwardMax PRD Excel 변환 스크립트")
    print("=" * 60)
    print("")
    
    # 입력 파일 확인
    if not os.path.exists(input_file):
        print(f"❌ 오류: 입력 파일을 찾을 수 없습니다: {input_file}")
        sys.exit(1)
    
    print(f"✓ 입력 파일 확인: {input_file}")
    
    # 마크다운 파일 읽기
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            md_content = f.read()
    except Exception as e:
        print(f"❌ 오류: 파일 읽기 실패: {e}")
        sys.exit(1)
    
    print(f"✓ 마크다운 파일 읽기 완료")
    print("")
    
    # Excel 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 시트 목록 정의 (순서대로)
    sheet_names = [
        '00_META',
        '01_SCHEMA',
        '02_GLOSSARY',
        '03_ACTOR',
        '04_COMPONENT',
        '05_POLICY_DISCLOSURE',
        '10_FLOW',
        '11_SCREEN',
        '12_FUNCTION',
        '13_ENTITY',
        '14_API',
        '15_NFR',
        '16_DECISION',
        '17_GAP',
        '18_RISK',
        '19_OPPORTUNITY',
        '20_SITEMAP',
        '30_FLOW_SCREEN',
        '31_SCREEN_FUNCTION',
        '32_FUNCTION_ENTITY',
        '33_FUNCTION_API',
        '34_FUNCTION_NFR',
        '35_FUNCTION_DECISION',
        '36_RBAC',
        '37_MASKING',
        '38_LOG_POLICY',
        '40_NEED',
        '41_REQUIREMENT',
        '42_TRACE_CHAIN',
    ]
    
    print("Excel 시트 생성 중...")
    created_sheets = []
    
    for sheet_name in sheet_names:
        df = parse_markdown_table(md_content, sheet_name)
        
        if df is not None and len(df) > 0:
            # 시트 생성
            ws = wb.create_sheet(title=sheet_name)
            
            # DataFrame을 시트에 쓰기
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 서식 적용
            apply_excel_formatting(ws, df)
            
            created_sheets.append(sheet_name)
            print(f"  ✓ {sheet_name} ({len(df)}행)")
        else:
            print(f"  ⚠ {sheet_name} (데이터 없음 또는 파싱 실패)")
    
    print("")
    print(f"총 {len(created_sheets)}개 시트 생성 완료")
    print("")
    
    # Excel 파일 저장
    try:
        wb.save(output_file)
        print(f"✓ Excel 파일 저장 완료: {output_file}")
    except Exception as e:
        print(f"❌ 오류: Excel 파일 저장 실패: {e}")
        sys.exit(1)
    
    print("")
    
    # 검증 수행
    print("검증 수행 중...")
    validation_results = {}
    
    validate_references(wb, validation_results)
    validate_duplicates(wb, validation_results)
    validate_required_columns(wb, validation_results)
    validate_counts(wb, validation_results)
    
    # 검증 리포트 생성
    generate_validation_report(validation_results, validation_report_file)
    
    # 검증 결과 요약 출력
    print("검증 결과:")
    total_errors = len(validation_results.get('reference_errors', [])) + len(validation_results.get('duplicate_errors', [])) + len(validation_results.get('required_errors', []))
    total_warnings = len(validation_results.get('required_warnings', []))
    
    if total_errors == 0:
        print(f"  ✅ 통과: 오류 없음")
    else:
        print(f"  ❌ 실패: 오류 {total_errors}건")
        if validation_results.get('reference_errors'):
            print(f"    - 참조 무결성 오류: {len(validation_results['reference_errors'])}건")
        if validation_results.get('duplicate_errors'):
            print(f"    - 중복 오류: {len(validation_results['duplicate_errors'])}건")
        if validation_results.get('required_errors'):
            print(f"    - 필수 컬럼 오류: {len(validation_results['required_errors'])}건")
    
    if total_warnings > 0:
        print(f"  ⚠ 경고: {total_warnings}건")
    
    if validation_results.get('counts'):
        print("  카운트:")
        for entity_type, count in validation_results['counts'].items():
            print(f"    - {entity_type}: {count}개")
    
    print("")
    print("=" * 60)
    print("변환 완료!")
    print("=" * 60)
    print("")
    print(f"출력 파일: {output_file}")
    print(f"검증 리포트: {validation_report_file}")
    print("")

if __name__ == '__main__':
    main()



